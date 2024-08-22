import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import PieChart, BarChart, Reference, label
import time

number_of_arg = 0

    
def calc_perc_polymorphic(polymorphic, no_markers, no_missing):
    return int(polymorphic)/(int(no_markers)-int(no_missing)) * 100

def calc_perc_het(no_het, no_markers, no_missing):
    return (int(no_het)/(int(no_markers) -int(no_missing))) * 100

def calc_perc_missing(missing, polymorphic):
    try:
        percentage_missing = (int(missing)/int(polymorphic)) * 100
    except ZeroDivisionError:
        percentage_missing = 0
        
    return percentage_missing

def calc_perc_outcross(outcross, polymorphic, no_missing):
    return (int(outcross)/(int(polymorphic) - int(no_missing))) * 100

def perc_hybridity(true, polymorphic, missing):
    return (int(true))/((int(polymorphic)-int(missing)))* 100
    
    
def hybridity(filename, saveas, min_missing_percentage=20, min_perc_polymorphic=20, min_perc_hybridity=50):
    max_perc_het = 20
    # start = time.time()
    # try:
    wb = load_workbook(filename, data_only=True)
    # except Exception as e:
    #     log.config(text=f"No such file or directory", fg="#800000")
    #     raise
    
    
    sheet = wb.worksheets[0]
    sheet.title = 'Results'
    
    dict_skip = {}
    # toskip = []
    marker_success = []
    # min_missing_percentage = 20
    # min_perc_polymorphic = 0
# parent_missing
    sheet.insert_cols(3, 11)
    # sheet.insert_cols(7, 3)
    
    undefined_missing_data = 0
    undefined_unpolymorphic_parent = 0
    undefined_parent_het = 0
    
    no_parent_polymophic = 0
    percent_polymorph = 0
    
    TRUE = 0
    FAILED = 0



    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    blue = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    pink = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    grey = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    polymophic_col = 2
    perc_polymorphic_col = 3
    
    no_parentHet_col = 4
    perc_parentHet_col = 5
    
    no_outcrossing_col = 6
    perc_outcrossing_col = 7
    
    trueF1_col = 8
    missing_col = 9
    perc_missing_col = 10
    hybridity_col = 11
    status_col = 12
    parent_column_col = 13
    marker_start_col = 14
    
    
    no_markers = sheet.max_column - marker_start_col
    
    def color_previous(col_number):
        sheet[f'{value[col_number].column_letter}{value[col_number].row - 1}'].fill = grey

    
    track_parent = 0
    track_parent_on_F1 = 0
    for k, (value, value2) in enumerate(zip(sheet.iter_rows(min_row=2, max_row=sheet.max_row),
                                                sheet.iter_rows(min_row=3, max_row=sheet.max_row)), 1):
        if value[parent_column_col].value == "Parent" and value2[parent_column_col].value == 'Parent':
            track_parent+=1
            no_parent_polymophic = 0
            parent_missing = 0
            
            parent1_het = 0
            parent2_het = 0

            parent1_missing = 0
            parent2_missing = 0
            value[parent_column_col].fill = green
            value2[parent_column_col].fill = green
            for i in range(marker_start_col, sheet.max_column):
                z = i - marker_start_col
                if len(marker_success) < no_markers:
                    marker_success.append([z, value[i].column_letter, 0])

                if value[i].value != value2[i].value and all(v[i].value not in ['Uncallable', '?'] for v in [value, value2]):
                    heterozyte = ['A:T', 'A:C', 'A:G', 'T:A', 'T:C', 'T:G', 'C:A', 'C:T', 'C:G', 'G:A', 'G:T', 'G:C',]
                    # if any(value.value in  
                    if value[i].value in heterozyte: 
                        value[i].fill = blue
                        value2[i].fill = blue
                        dict_skip[f"parent{track_parent}{value[i].column_letter}"] = "Skip"
                        parent1_het += 1
                    elif value2[i].value in heterozyte:
                        value[i].fill = blue
                        value2[i].fill = blue
                        dict_skip[f"parent{track_parent}{value[i].column_letter}"] = "Skip"
                        parent2_het += 1
                       
                    else:
                        value[i].fill = green
                        value2[i].fill = green
                        exp3 = '{}:{}'.format(str(value[i].value)[-1], str(value2[i].value)[-1])
                        exp4 = '{}:{}'.format(str(value2[i].value)[-1], str(value[i].value)[-1])
                        
                        dict_skip[f"parent{track_parent}{value[i].column_letter}"] = [exp3, exp4]
                        
                    no_parent_polymophic += 1
                    marker_success[z] = ([z, value[i].column_letter, ((marker_success[z])[2] + 1)])
                    
                elif value[i].value == value2[i].value and any(v[i].value in ['Uncallable', '?'] for v in [value, value2]):
                    dict_skip[f"parent{track_parent}{value[i].column_letter}"] = "Skip"
                    value[i].fill = red
                    value2[i].fill = red
                    parent1_missing += 1
                    parent2_missing += 1
                    parent_missing += 1
                    
                else:
                    dict_skip[f"parent{track_parent}{value[i].column_letter}"] = "Skip"
                    

                    if all(v.value not in ['G:G', 'A:A', 'C:C', 'T:T', 'Uncallable', '?',] for v in [value[i], value2[i]]):
                        # print('hetero')
                        
                        # print(value[i].coordinate)
                        value[i].fill = blue
                        value2[i].fill = blue
                        parent1_het += 1
                        parent2_het += 1
                        
                    elif any(v[i].value in ['Uncallable', '?'] for v in [value, value2]):
                        value[i].fill = red
                        value2[i].fill = red
                        parent1_missing += 1
                        parent2_missing += 1
                        parent_missing += 1
    
                        
                    elif value2[i].value in ['Uncallable', '?']:
                        value[i].fill = red
                        value2[i].fill = red
                        parent2_missing += 1
                        
                    else:
                        value[i].fill = red
                        value2[i].fill = red

            value[missing_col].value = parent_missing
            value[missing_col].fill = grey
            value[perc_missing_col].value = calc_perc_missing(parent_missing, no_parent_polymophic)
            value[perc_missing_col].fill = grey
            
            value2[missing_col].value = parent_missing
            value2[missing_col].fill = grey
            value2[perc_missing_col].value = calc_perc_missing(parent_missing, no_parent_polymophic)
            value2[perc_missing_col].fill = grey
            
            
            # print(parent1_missing, calc_perc_missing(parent1_missing, parent_polymophic), 'parent2: ', parent2_missing, calc_perc_missing(parent2_missing, parent_polymophic), "parent_poly: ", parent_polymophic)
            
            value[polymophic_col].value = no_parent_polymophic
            value[polymophic_col].fill = grey
            
            value2[polymophic_col].value = value[polymophic_col].value
            value2[polymophic_col].fill = grey
            
            value[no_parentHet_col].value = parent1_het
            value2[no_parentHet_col].value = parent2_het
            
            value[no_parentHet_col].fill = grey
            value2[no_parentHet_col].fill = grey
            
            perc_het = (calc_perc_het(value[no_parentHet_col].value, no_markers, value[missing_col].value))
            perc_het2 = (calc_perc_het(value2[no_parentHet_col].value, no_markers, value2[missing_col].value))
            
            value[perc_parentHet_col].value = perc_het
            value2[perc_parentHet_col].value = perc_het2
            
            value[perc_parentHet_col].fill = grey
            value2[perc_parentHet_col].fill = grey
            
            
            # print(value[polymophic_col].value, no_markers, value[polymophic_col].coordinate)
            percent_polymorph = calc_perc_polymorphic(value[polymophic_col].value, no_markers, value[perc_missing_col].value)
            
            
            value[perc_polymorphic_col].value = calc_perc_polymorphic(value[polymophic_col].value, no_markers, value[missing_col].value)
            value[perc_polymorphic_col].fill = grey
            
            value2[perc_polymorphic_col].value = calc_perc_polymorphic(value2[polymophic_col].value, no_markers, value2[missing_col].value)
            value2[perc_polymorphic_col].fill = grey

    # print([*dict_skip.get(f"parent{track_parent}{value[i].column_letter}")])

    for k, (value, value2) in enumerate(zip(sheet.iter_rows(min_row=2, max_row=sheet.max_row),
                                                sheet.iter_rows(min_row=3, max_row=sheet.max_row)), 1):
        if (value[parent_column_col].value == 'Parent' and value2[parent_column_col].value == 'F1') or (value[parent_column_col].value == 'F1' and value2[parent_column_col].value == 'F1'):
            value2[polymophic_col].fill = grey
            value2[no_parentHet_col].fill = grey
            value2[perc_parentHet_col].fill = grey
            
            no_outcross = 0
            missiing = 0
            true = 0
            if (value[parent_column_col].value == 'Parent' and value2[parent_column_col].value == 'F1'):
                
                percent_polymorph = int(value[perc_polymorphic_col].value)
                no_parent_polymophic = int(value[polymophic_col].value)
                track_parent_on_F1 += 1
                
                first_parent = f'{value[perc_parentHet_col].column_letter}{value[perc_parentHet_col].row - 1}'
                
                perc_het = int(value[perc_parentHet_col].value) if int(value[perc_parentHet_col].value) > sheet[first_parent].value else sheet[first_parent].value
                
              

            for i in range(marker_start_col, sheet.max_column):
                test_skip = f"parent{track_parent_on_F1}{value[i].column_letter}"
                dict_skip.get(test_skip)

                "Using dictionary"
                if value2[i].value not in ['G:G', 'A:A', 'C:C', 'T:T', 'Uncallable', '?',] and value2[i].value in dict_skip.get(test_skip):
                    value2[i].fill = orange
                    true += 1
                else:
                    if value2[i].value in ['Uncallable', '?'] and dict_skip.get(test_skip) not in ['Skip']:
                        missiing += 1
                    
                    elif value2[i].value not in ['G:G', 'A:A', 'C:C', 'T:T', dict_skip.get(test_skip)] and dict_skip.get(test_skip) not in ['Skip', None]:
                        value2[i].fill = pink
                        no_outcross += 1
                        

            value2[missing_col].value = missiing
            value2[trueF1_col].value = true
            value2[no_outcrossing_col].value = no_outcross
            
            # color 
            value2[missing_col].fill = grey
            value2[trueF1_col].fill = grey
            value2[no_outcrossing_col].fill = grey
            
            
            
            if value[missing_col].value == None:
                value[missing_col].fill = grey
                
                color_previous(missing_col)
                
                
            if value[trueF1_col].value == None:
                value[trueF1_col].fill = grey
                
                color_previous(trueF1_col)


            if value[no_outcrossing_col].value == None:
                value[no_outcrossing_col].fill = grey
                
                color_previous(no_outcrossing_col)
            
            try:
                percent_missing = calc_perc_missing(missiing, no_parent_polymophic)
            except ZeroDivisionError:
                percent_missing = 999

            try:
                percent_outcross = calc_perc_outcross(no_outcross, no_parent_polymophic, missiing)
            except ZeroDivisionError:
                print(no_outcross, no_parent_polymophic, missiing)
                print('=========>', value2[i].coordinate)
                percent_outcross = 0
            

            #color
            value2[perc_polymorphic_col].fill = grey
            
            if value[perc_polymorphic_col].value == None:
                value[perc_polymorphic_col].fill = grey
                color_previous(perc_polymorphic_col)
                
            # if value[perc_polymorphic].value == None:
            #     value[perc_polymorphic].fill = grey
            #     color_previous(perc_polymorphic)
            
            if percent_missing == 999:
                value2[perc_missing_col].value = 'NA'
            else:
                value2[perc_missing_col].value = percent_missing
                
            #color
            value2[perc_missing_col].fill = grey
            
            value2[perc_outcrossing_col].value = percent_outcross
            value2[perc_outcrossing_col].fill = grey
            
            if value[perc_missing_col].value == None:
                value[perc_missing_col].fill = grey
                color_previous(perc_missing_col)
            
            if value[perc_outcrossing_col].value == None:
                value[perc_outcrossing_col].fill = grey
                color_previous(perc_outcrossing_col)


            if (percent_polymorph <= min_perc_polymorphic):
                value2[hybridity_col].value = 'NA'
                value2[status_col].value = 'Undetermine: Parent not polymorphic'
                undefined_unpolymorphic_parent += 1
            
            if (((percent_missing >= min_missing_percentage) or (percent_missing == 999)) and value2[status_col].value is None):
                value2[hybridity_col].value = 'NA'
                value2[status_col].value = 'Undetermine: missing data'
                undefined_missing_data += 1
                
            # if (perc_het > max_perc_het):
            #     value2[hybridity_col].value = 'NA'
            #     value2[status_col].value = 'Undetermine: Parent heterogyzote'
            #     undefined_parent_het += 1
            
            value2[hybridity_col].fill = grey
            value2[status_col].fill = grey
            
            
            if value[hybridity_col].value == None:
                value[hybridity_col].fill = grey
                color_previous(hybridity_col)
            if value[status_col].value == None:
                value[status_col].fill = grey
                color_previous(status_col)
            try:
                if value2[hybridity_col].value is None:
                    percent_hybridity = perc_hybridity(true, no_parent_polymophic, missiing)
                    
                    if percent_hybridity < min_perc_hybridity and value2[status_col].value is None:
                        if perc_het > max_perc_het:
                            undefined_parent_het += 1
                            
                            value2[hybridity_col].value = 'NA'
                            value2[status_col].value = 'Undetermine: Parent Heterozygous'
                            # print(perc_het, perc_het2, percent_hybridity)
                        else:
                            value2[status_col].value = 'SELF'
                            FAILED += 1
                    else:
                        value2[status_col].value = 'TRUE CROSS'
                        TRUE += 1
                
                    if value2[hybridity_col].value is None:
                        value2[hybridity_col].value = percent_hybridity
                    value2[hybridity_col].fill = grey
                    
                    
                    if value[hybridity_col].value == None:
                        value[hybridity_col].fill = grey
                        color_previous(hybridity_col)
                    
        
            except ValueError:
                pass
            except ZeroDivisionError:
                value2[hybridity_col].value = 'NA'
                value2[hybridity_col].fill = grey
                if value[hybridity_col].value == None:
                    value[hybridity_col].fill = grey
                    color_previous(hybridity_col)
        else:
            continue 
        

    sheet['C1'].value = '#polymorphic'
    sheet['D1'].value = '%polymorphic'
    sheet['E1'].value = '#parentHet'
    sheet['F1'].value = '%parentHet'
    sheet['G1'].value = '#NonParentAllele'
    sheet['H1'].value = '%NonParentAllele'
    sheet['I1'].value = '#true'
    sheet['J1'].value = '#missing'
    sheet['K1'].value = '%missing'
    sheet['L1'].value = '%hybridity'
    sheet['M1'].value = 'Status'
    
    
    
    sheet['C1'].fill = grey
    sheet['D1'].fill = grey
    sheet['E1'].fill = grey
    sheet['F1'].fill = grey
    sheet['G1'].fill = grey
    sheet['H1'].fill = grey
    sheet['I1'].fill = grey
    sheet['J1'].fill = grey 
    sheet['K1'].fill = grey 
    sheet['L1'].fill = grey
    sheet['M1'].fill = grey 
     # wb.save('compiled_hybridity.xlsx')
    # print(TRUE, FAILED, undefined_missing_data, undefined_unpolymorphic_parent)
    TOTAL = TRUE + FAILED + undefined_missing_data + undefined_unpolymorphic_parent + undefined_parent_het
    
    
    data = [
        ["Status", 'Value'],
        ["TRUE CROSS", TRUE],
        ["SELF", FAILED],
        ["Undetermined: Parent not polymorphic", undefined_unpolymorphic_parent],
        ["Undetermined: Missing data", undefined_missing_data],
        ["Undetermined: Parent Heterozygous", undefined_parent_het],
        ["TOTAL", TOTAL],
    ]

    chartSheet = wb.create_sheet(title='Hybridity')
    for row in data:
        chartSheet.append(row)
    
    pie_chart = PieChart()
    data = Reference(chartSheet, min_col=2, min_row=1, max_col=2, max_row=len(data)-1)
    categories = Reference(chartSheet, min_col=1, min_row=2, max_row=len(data))

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)
    pie_chart.title = "HYBRIDITY PIE CHART"
    
    # Add data labels to the chart
    pie_chart.dataLabels = label.DataLabelList()
    pie_chart.dataLabels.showPercent = True  # Display percentage
    # pie_chart.dataLabels.showVal = True  # Display actual values
    
    pie_chart.dataLabels.position = "bestFit"  # Position labels at best places
    pie_chart.dataLabels.number_format = "0.0%"  # Format percentages

    
    chartSheet.add_chart(pie_chart, "H8")
    
    
    barData = [
            #    ["SNPs", 'Total Parents', "No of Polymorphic Parent", 'Frequenxy of Success'],
               [sheet[f'{i[1]}1'].value, track_parent, i[2], ((i[2]/track_parent)*100)] for i in marker_success
               ]
    barData.insert(0, ["SNPs", '#Parent Combination', "Polymorphism Frequency", 'Marker Efficiency (%)'])
    barSheet = wb.create_sheet(title='Efficiency')
    
    for row in barData:
        barSheet.append(row)
        
        
    barchart = BarChart()
    barchart.title = "SNPs Performance"
    barchart.x_axis.title = "Markers"
    barchart.y_axis.title = "Marker Efficiency"
    
    data = Reference(barSheet, min_col=4, min_row=1, max_col=4, max_row=len(barData))
    labels = Reference(barSheet, min_col=1, min_row=2, max_row=len(barData))

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(labels)
    barchart.legend = None
    
    barSheet.add_chart(barchart, "H8")
    
    wb.save(saveas)
    # stop = time.time()
    # print(stop-start)
    

if __name__ == "__main__":
    # print(len(sys.argv))
    if len(sys.argv) not in [3, 6]:
        print(
            """You need to specify all the arguments
            See Guideline Below
            first argument -- the file you want to analyse
            second argument -- where you want the file to be saved with file name
            third argument -- the maximum missing percentage acceptable
            forth argument -- the minimum polymorphic percentage acceptable
            fifth argument -- the minimum hybridity percentage acceptable
            You may specify just two arguments if you want to use the default thresholds that is %missing = 20, %polymorphic=20 and %hybridity=50
            
            See the github repo for more understanding
            """
        )

        sys.exit(1)
    else:
        number_of_arg = len(sys.argv)


if number_of_arg >= 3:
    filename = str(sys.argv[1])
    saveas = str(sys.argv[2])
    hybridity(filename, saveas)
if number_of_arg == 6:
    min_missing_percentage = int(sys.argv[3]) if sys.argv[3] is None else 20
    min_perc_polymorphic = int(sys.argv[4]) if sys.argv[4] is None else 20
    min_perc_hybridity = int(sys.argv[5]) if sys.argv[5] is None else 50
    hybridity(filename, saveas, min_missing_percentage, min_perc_polymorphic, min_perc_hybridity)

        