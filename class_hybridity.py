from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import PieChart, BarChart, Reference, label


class CalculationLogics():
    def calc_perc_polymorphic(polymorphic, no_markers):
        return (int(polymorphic)/int(no_markers)) * 100

    def calc_perc_het(no_het, no_markers):
        return (int(no_het)/int(no_markers)) * 100

    def calc_perc_missing(missing, polymorphic):
        return (int(missing)/int(polymorphic)) * 100

    def calc_perc_outcross(outcross, polymorphic):
        return (int(outcross)/int(polymorphic)) * 100

    def perc_hybridity(true, polymorphic, missing):
        return (int(true))/(int(str(polymorphic))-int(str(missing)))* 100


class HybridQC:
    def __init__(self, filename : str, saveas, min_missing_percentage=20, min_perc_polymorphic=20, min_perc_hybridity=50):
        self.file = filename
        self.min_missing_percentage= min_missing_percentage
        self.min_perc_polymorphic= min_perc_polymorphic
        self.min_perc_hybridity= min_perc_hybridity
        self.saveas = saveas
        self.green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        self.blue = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        self.red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        self.pink = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        self.grey = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
       
        
        self.wb = load_workbook(self.file, data_only=True)
        self.sheet = self.wb.worksheets[0]
        self.max_perc_het = 20
        self.sheet.title = 'Results'
        self.dict_skip = {}
        # self.# toskip = []
        self.marker_success = []
        # self.# min_missing_percentage = 20
        # self.# min_perc_polymorphic = 0

        self.sheet.insert_cols(3, 11)
        # self.# sheet.insert_cols(7, 3)

        self.undefined_missing_data = 0
        self.undefined_unpolymorphic_parent = 0
        self.undefined_parent_het = 0

        self.parent_polymophic = 0
        self.percent_polymorph = 0

        self.TRUE = 0
        self.FAILED = 0


        self.polymophic_col = 2
        self.perc_polymorphic_col = 3

        self.no_parentHet_col = 4
        self.perc_parentHet_col = 5

        self.no_outcrossing_col = 6
        self.perc_outcrossing_col = 7

        self.trueF1_col = 8
        self.missing_col = 9
        self.perc_missing_col = 10
        self.hybridity_col = 11
        self.status_col = 12
        self.parent_col = 13
        self.marker_start_col = 14
        self.no_markers = self.sheet.max_column - self.marker_start_col
        self.track_parent = 0
        self.track_parent_on_F1 = 0
        
        self.heterozyte = ['A:T', 'A:C', 'A:G', 'T:A', 'T:C', 'T:G', 'C:A', 'C:T', 'C:G', 'G:A', 'G:T', 'G:C',]

    def start(self):
        self.checkPolymorphicParent()
        self.f1check()
        self.createStatHeaders()
        self.createPieChart()
        self.createBarChart()
        self.save("done"+self.file)
        
    def checkPolymorphicParent(self):
        for k, (value, value2) in enumerate(zip(self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row),
                                                self.sheet.iter_rows(min_row=3, max_row=self.sheet.max_row)), 1):
            
            self.curr_parent = value
            self.next_parent = value2
            
            curr_parent =  self.curr_parent
            next_parent =  self.next_parent
            if self.parentPair():
                self.track_parent+=1
                self.parent_polymophic = 0
                self.parent_missing = 0
                
                self.parent1_het = 0
                self.parent2_het = 0
                
                self.parent1_missing = 0
                self.parent2_missing = 0
                
                curr_parent[self.parent_col].fill = self.green
                next_parent[self.parent_col].fill = self.green
                
                for i in range(self.marker_start_col, self.sheet.max_column):
                    marker_index = i - self.marker_start_col
                    if len(self.marker_success) < self.no_markers:
                                            #marker_no  #marker_col                #initialize number of success
                        self.marker_success.append([marker_index, curr_parent[i].column_letter, 0])

                    if curr_parent[i].value != next_parent[i].value and self.isNotMissing(i):
                        self.colorAndSkipIfParentHet(self.parent1_het, self.parent2_het, i)
                        self.marker_success[marker_index] = ([marker_index, value[i].column_letter, ((self.marker_success[marker_index])[2] + 1)])
                    
                    elif self.bothMissing(i):
                        self.dict_skip[f"parent{self.track_parent}{self.curr_parent[i].column_letter}"] = "Skip"
                        self.curr_parent[i].fill = self.red
                        self.next_parent[i].fill = self.red
                        self.parent1_missing += 1
                        self.parent2_missing += 1
                        self.parent_missing += 1
                        
                    else:
                    
                        self.dict_skip[f"parent{self.track_parent}{self.curr_parent[i].column_letter}"] = "Skip"
                        

                        if all(v.value not in ['G:G', 'A:A', 'C:C', 'T:T', 'Uncallable', '?',] for v in [self.curr_parent[i], self.next_parent[i]]):
                            self.curr_parent[i].fill = self.blue
                            self.next_parent[i].fill = self.blue
                            self.parent1_het += 1
                            self.parent2_het += 1

                        else: 
                            self.curr_parent[i].fill = self.red
                            self.next_parent[i].fill = self.red
                            
                            
                self.setPolymorphicHybridityValues()
                self.colorParentGrey()
                # print("checking D3:", self.sheet["D3"].value)
        # print("Checking parents done")
    
    def set_polymorphic_columns(self,):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        
        curr_parent[self.polymophic_col].value = self.parent_polymophic
        next_parent[self.polymophic_col].value = self.parent_polymophic
        
        perc_poly =  (self.parent_polymophic/int(self.no_markers)) * 100
        curr_parent[self.perc_polymorphic_col].value = perc_poly
        next_parent[self.perc_polymorphic_col].value = perc_poly
        # print("fisrt: ",perc_poly, curr_parent[self.perc_polymorphic_col].coordinate)
        # print(next_parent[self.perc_polymorphic_col].value)

    def set_heterozygote_columns(self):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        
        curr_parent[self.no_parentHet_col].value = self.parent1_het
        next_parent[self.no_parentHet_col].value = self.parent2_het
        
        perc_het1 =  (self.parent1_het/int(self.no_markers)) * 100
        perc_het2 =  (self.parent2_het/int(self.no_markers)) * 100
        curr_parent[self.perc_parentHet_col].value = perc_het1
        next_parent[self.perc_parentHet_col].value = perc_het2
        
        # return (int(self.parent1_het)/int(self.no_markers)) * 100

    def set_missing_columns(self,):
        self.curr_parent[self.missing_col].value = self.parent_missing
        self.next_parent[self.missing_col].value = self.parent_missing
        self.curr_parent[self.missing_col].fill = self.grey
        self.next_parent[self.missing_col].fill = self.grey
        perc_missing = (int(self.parent_missing)/int(self.parent_polymophic)) * 100
        self.curr_parent[self.perc_missing_col].value = perc_missing
        self.next_parent[self.perc_missing_col].value = perc_missing
        # return (int(self.parent_missing)/int(self.parent_polymophic)) * 100

    def calc_perc_outcross(self):
        outcross = self.no_outcross
        polymorphic = self.parent_polymophic
        no_missing = self.missiing
        return (int(outcross)/(int(polymorphic) - int(no_missing))) * 100

    def perc_hybridity(self,):
        true = self.true
        polymorphic = self.parent_polymophic
        missing = self.missiing
        return (int(true))/(int((polymorphic))-int((missing)))* 100
    
    def calc_perc_het(no_het, no_markers, no_missing):
        return (int(no_het)/(int(no_markers) -int(no_missing))) * 100

    def calc_perc_polymorphic(polymorphic, no_markers, no_missing):
        return int(polymorphic)/(int(no_markers)-int(no_missing)) * 100

    def calc_perc_missing(self):
        try:
            percentage_missing = (int(self.missiing)/int(self.parent_polymophic)) * 100
        except ZeroDivisionError:
            percentage_missing = 0
            
        return percentage_missing
 
    def setPolymorphicHybridityValues(self):
        self.set_heterozygote_columns()
        self.set_polymorphic_columns()
        self.set_missing_columns()
        
        # curr_parent[self.polymophic_col].value = self.parent_polymophic
        # next_parent[self.polymophic_col].value = curr_parent[self.polymophic_col].value
        # curr_parent[self.no_parentHet_col].value = self.parent1_het
        # next_parent[self.no_parentHet_col].value = self.parent2_het
        
    def colorAndSkipIfParentHet(self, parent1_het, parent2_het, i):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        
        if curr_parent[i].value in self.heterozyte: 
            curr_parent[i].fill = self.blue
            next_parent[i].fill = self.blue
            self.dict_skip[f"parent{self.track_parent}{curr_parent[i].column_letter}"] = "Skip"
            self.parent1_het += 1
        elif next_parent[i].value in self.heterozyte:
            curr_parent[i].fill = self.blue
            next_parent[i].fill = self.blue
            self.dict_skip[f"parent{self.track_parent}{curr_parent[i].column_letter}"] = "Skip"
            self.parent2_het += 1
                        
        else:
            curr_parent[i].fill = self.green
            next_parent[i].fill = self.green
            exp3 = '{}:{}'.format(str(curr_parent[i].value)[-1], str(next_parent[i].value)[-1])
            exp4 = '{}:{}'.format(str(next_parent[i].value)[-1], str(curr_parent[i].value)[-1])
                            
            self.dict_skip[f"parent{self.track_parent}{curr_parent[i].column_letter}"] = [exp3, exp4]
        self.parent_polymophic += 1

    def colorParentGrey(self):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        grey = self.grey
        
        curr_parent[self.polymophic_col].fill = grey
        next_parent[self.polymophic_col].fill = grey
        curr_parent[self.no_parentHet_col].fill = grey
        next_parent[self.no_parentHet_col].fill = grey
        curr_parent[self.perc_parentHet_col].fill = grey
        next_parent[self.perc_parentHet_col].fill = grey
        curr_parent[self.perc_polymorphic_col].fill = grey
        next_parent[self.perc_polymorphic_col].fill = grey
        curr_parent[self.perc_missing_col].fill = grey
        next_parent[self.perc_missing_col].fill = grey
        
    def isNotMissing(self, i):
        return all(v[i].value not in ['Uncallable', '?'] for v in [self.curr_parent, self.next_parent])
    
    def bothMissing(self, i):
        value = self.curr_parent
        value2 = self.next_parent
        value[i].value == value2[i].value and any(v[i].value in ['Uncallable', '?'] for v in [value, value2])
    
    def f1check(self):
        grey = self.grey
        
        for k, (value, value2) in enumerate(zip(self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row),
                                                    self.sheet.iter_rows(min_row=3, max_row=self.sheet.max_row)), 1):
            self.curr_parent = value
            self.next_parent = value2
            
            if self.f1StartandContinues():
                value2[self.polymophic_col].fill = grey
                value2[self.no_parentHet_col].fill = grey
                value2[self.perc_parentHet_col].fill = grey
                
                self.no_outcross = 0
                self.missiing = 0
                self.true = 0
                
                if (value[self.parent_col].value == 'Parent' and value2[self.parent_col].value == 'F1'):
                    self.percent_polymorph = int(value[self.perc_polymorphic_col].value)
                    self.parent_polymophic = int(value[self.polymophic_col].value)
                    self.track_parent_on_F1 += 1
                    
                    first_parent = f'{value[self.perc_parentHet_col].column_letter}{value[self.perc_parentHet_col].row - 1}'
                
                    self.perc_het = int(value[self.perc_parentHet_col].value) if int(value[self.perc_parentHet_col].value) > self.sheet[first_parent].value else self.sheet[first_parent].value
                
                self.determineF1hybridity()
                self.setF1Stat()
                # colorself. 
                self.colorF1Stat()
                #color
                # if value[perc_polymorphic].value == None:
                #     value[perc_polymorphic].fill = grey
                #     color_previous(perc_polymorphic)
  
                #color
                self.setUndetermineF1()
                    
                # if (perc_parentHet >= min_perc_het):
                #     value2[self.hybridity_col].value = 'NA'
                #     value2[status_col].value = 'Undetermine: Parent heterogyzote'
                #     undefined_parent_het += 1

                    
                self.determineHybridity()
            else:
                continue

    def determineHybridity(self):
        grey = self.grey
        try:
            if self.next_parent[self.hybridity_col].value is None:
                percent_hybridity = self.perc_hybridity()
                        
                if percent_hybridity < self.min_perc_hybridity and self.next_parent[self.status_col].value is None:
                    if self.perc_het > self.max_perc_het:
                        self.undefined_parent_het += 1
                        
                        self.next_parent[self.hybridity_col].value = 'NA'
                        self.next_parent[self.status_col].value = 'Undetermine: Parent Heterozygous'
                        # print(perc_het, perc_het2, percent_hybridity)
                    else:
                        self.next_parent[self.status_col].value = 'SELF'
                        self.FAILED += 1
                else:
                    self.next_parent[self.status_col].value = 'TRUE CROSS'
                    self.TRUE += 1
                #     self.next_parent[self.status_col].value = 'FAILED'
                #     self.FAILED += 1
                # else:
                #     self.next_parent[self.status_col].value = 'TRUE'
                #     self.TRUE += 1
                    

                self.next_parent[self.hybridity_col].value = percent_hybridity
                self.next_parent[self.hybridity_col].fill = grey
                        
        except ValueError:
            pass
        except ZeroDivisionError:
            self.next_parent[self.hybridity_col].value = 'NA'

    def setUndetermineF1(self):
        if (self.percent_polymorph <= self.min_perc_polymorphic):
            self.next_parent[self.hybridity_col].value = 'NA'
            self.next_parent[self.status_col].value = 'Undetermine: Parent not polymorphic'
            self.undefined_unpolymorphic_parent += 1
                
        if (((self.next_parent[self.perc_missing_col].value >= self.min_missing_percentage) or (self.next_parent[self.perc_missing_col].value == 999)) and self.next_parent[self.status_col].value is None):
            self.next_parent[self.hybridity_col].value = 'NA'
            self.next_parent[self.status_col].value = 'Undetermine: missing data'
            self.undefined_missing_data += 1

    def colorF1Stat(self):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        grey = self.grey
        cells = [self.missing_col, self.hybridity_col, self.trueF1_col, self.no_outcrossing_col, self.hybridity_col, self.status_col,
                 self.perc_polymorphic_col, self.perc_missing_col, self.perc_outcrossing_col, ]
        
        for i in cells: 
            next_parent[i].fill = grey
            
            if curr_parent[i].value == None:
                curr_parent[i].fill = grey
                self.color_previous(i)

    def setF1Stat(self):
        next_parent =  self.next_parent
        next_parent[self.missing_col].value = self.missiing
        next_parent[self.no_outcrossing_col].value = self.no_outcross
        next_parent[self.trueF1_col].value = self.true
        
                
                # next_parent[self.perc_outcrossing].value = percent_outcross
        try:
            next_parent[self.perc_missing_col].value = self.calc_perc_missing()
        except ZeroDivisionError:
            next_parent[self.perc_missing_col].value = 'NA'

                # if percent_missing != 999:
                #     next_parent[self.perc_missing].value = percent_missing
                # else:
                #     next_parent[self.perc_missing].value = 'NA'

        try:
            next_parent[self.perc_outcrossing_col].value = self.calc_perc_outcross()
        except ZeroDivisionError:
            next_parent[self.perc_outcrossing_col].value = 0
            # next_parent[self.perc_outcrossing_col].value = 'NA'

    def determineF1hybridity(self):
        orange = self.orange
        pink = self.pink
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        
        # print(next_parent)
        
        for i in range(self.marker_start_col, self.sheet.max_column):
            self.test_skip = f"parent{self.track_parent_on_F1}{curr_parent[i].column_letter}"

            self.dict_skip.get(self.test_skip)

            "Using dictionary"
            if next_parent[i].value not in ['G:G', 'A:A', 'C:C', 'T:T', 'Uncallable', '?',] and next_parent[i].value in self.dict_skip.get(self.test_skip):
                next_parent[i].fill = orange
                self.true += 1
            else:
                if next_parent[i].value in ['Uncallable', '?'] and self.dict_skip.get(self.test_skip) not in ['Skip']:
                    self.missiing += 1
                        
                elif next_parent[i].value not in ['G:G', 'A:A', 'C:C', 'T:T', self.dict_skip.get(self.test_skip)] and self.dict_skip.get(self.test_skip) not in ['Skip', None]:
                    next_parent[i].fill = pink
                    self.no_outcross += 1

    def f1StartandContinues(self):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        return (curr_parent[self.parent_col].value == 'Parent' and next_parent[self.parent_col].value == 'F1') or (curr_parent[self.parent_col].value == 'F1' and next_parent[self.parent_col].value == 'F1')


    def parentPair(self):
        curr_parent =  self.curr_parent
        next_parent =  self.next_parent
        return curr_parent[self.parent_col].value == "Parent" and next_parent[self.parent_col].value == 'Parent'
    
    def createStatHeaders(self):
        grey = self.grey
        sheet = self.sheet
        sheet['C1'].value = '#polymorphic'
        sheet['D1'].value = '%polymorphic'
        sheet['E1'].value = '#parentHet'
        sheet['F1'].value = '%parentHet'
        sheet['G1'].value = '#NonParentAllele'
        sheet['H1'].value = '%NonParentAllele'
        sheet['I1'].value = '#true'
        sheet['J1'].value = '#missing'
        sheet['K1'].value = '%missing'
        sheet['L1'].value = '%hybridity'
        sheet['M1'].value = 'Status'
        
        
        
        sheet['C1'].fill = grey
        sheet['D1'].fill = grey
        sheet['E1'].fill = grey
        sheet['F1'].fill = grey
        sheet['G1'].fill = grey
        sheet['H1'].fill = grey
        sheet['I1'].fill = grey
        sheet['J1'].fill = grey 
        sheet['K1'].fill = grey 
        sheet['L1'].fill = grey
        sheet['M1'].fill = grey
        
    def color_previous(self, col_number):
        curr_parent =  self.curr_parent
        # next_parent =  self.next_parent
        self.sheet[f'{curr_parent[col_number].column_letter}{curr_parent[col_number].row - 1}'].fill = self.grey
        
        
    def createPieChart(self):
        TRUE=self.TRUE 
        FAILED=self.FAILED 
        undefined_missing_data=self.undefined_missing_data 
        undefined_unpolymorphic_parent=self.undefined_unpolymorphic_parent 
        undefined_parent_het=self.undefined_parent_het
        wb = self.wb

        TOTAL = TRUE + FAILED + undefined_missing_data + undefined_unpolymorphic_parent + undefined_parent_het
    
    
        data = [
            ["Status", 'Value'],
            ["TRUE", self.TRUE],
            ["FAILED", FAILED],
            ["Undetermined: Parent not polymorphic", undefined_unpolymorphic_parent],
            ["Undetermined: Missing data", undefined_missing_data],
            # ["Undetermined: Parent Heterogyte", undefined_parent_het],
            ["TOTAL", TOTAL],
        ]

        chartSheet = wb.create_sheet(title='Hybridity')
        for row in data:
            chartSheet.append(row)
        
        pie_chart = PieChart()
        data = Reference(chartSheet, min_col=2, min_row=1, max_col=2, max_row=len(data)-1)
        categories = Reference(chartSheet, min_col=1, min_row=2, max_row=len(data))

        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(categories)
        pie_chart.title = "HYBRIDITY PIE CHART"
        
        # Add data labels to the chart
        pie_chart.dataLabels = label.DataLabelList()
        pie_chart.dataLabels.showPercent = True  # Display percentage
        # pie_chart.dataLabels.showVal = True  # Display actual values
        
        pie_chart.dataLabels.position = "bestFit"  # Position labels at best places
        pie_chart.dataLabels.number_format = "0.0%"  # Format percentages

        
        chartSheet.add_chart(pie_chart, "H8")
    def createBarChart(self):
        wb = self.wb
        barData = [
        #    ["SNPs", 'Total Parents', "No of Polymorphic Parent", 'Frequenxy of Success'],
            [self.sheet[f'{i[1]}1'].value, self.track_parent, i[2], ((i[2]/self.track_parent)*100)] for i in self.marker_success
            ]
        barData.insert(0, ["SNPs", '#Parent Combination', "Polymorphism Frequency", 'Marker Efficiency (%)'])
        barSheet = wb.create_sheet(title='Efficiency')
        
        for row in barData:
            barSheet.append(row)
            
            
        barchart = BarChart()
        barchart.title = "SNPs Performance"
        barchart.x_axis.title = "Markers"
        barchart.y_axis.title = "Marker Efficiency"
        
        data = Reference(barSheet, min_col=4, min_row=1, max_col=4, max_row=len(barData))
        labels = Reference(barSheet, min_col=1, min_row=2, max_row=len(barData))

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(labels)
        barchart.legend = None
        
        barSheet.add_chart(barchart, "H8")
        
    def save(self, saveas):
        self.wb.save(saveas)

if __name__ == "main":
    hybrid = HybridQC("1000 F1 + parent + 22SNPS.xlsx")
    hybrid.start()